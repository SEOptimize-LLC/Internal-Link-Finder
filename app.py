#!/usr/bin/env python3
"""
Internal Link Finder - Streamlit Web App

A web-based tool for discovering internal linking opportunities using vector embeddings
from Screaming Frog, enhanced with Google Search Console performance metrics.

Based on Everett Sizemore's methodology:
https://moz.com/blog/internal-linking-opportunities-with-vector-embeddings

Deploy on Streamlit Cloud: https://streamlit.io/cloud
"""

import streamlit as st
import pandas as pd
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

# Page configuration
st.set_page_config(
    page_title="Internal Link Finder",
    page_icon="🔗",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ----------- COLUMN NAME MAPPINGS -----------

GSC_URL_COLUMNS = [
    'landing page', 'landing pages', 'url', 'urls', 'page', 'pages',
    'address', 'landing_page', 'landing_pages', 'top pages', 'top page'
]

GSC_CLICKS_COLUMNS = [
    'clicks', 'click', 'total clicks', 'total_clicks'
]

GSC_IMPRESSIONS_COLUMNS = [
    'impressions', 'impression', 'total impressions', 'total_impressions', 'impr'
]

GSC_POSITION_COLUMNS = [
    'average position', 'avg. position', 'avg position', 'avg. pos', 'avg pos',
    'position', 'pos', 'average_position', 'avg_position', 'avg_pos',
    'mean position', 'mean_position'
]

GSC_CTR_COLUMNS = [
    'ctr', 'url ctr', 'average ctr', 'avg. ctr', 'avg ctr',
    'click through rate', 'click_through_rate', 'url_ctr', 'avg_ctr'
]


# ----------- HELPER FUNCTIONS -----------

def find_column(df, possible_names, required=True):
    """Find a column in the dataframe that matches any of the possible names."""
    df_columns_lower = {col.lower().strip(): col for col in df.columns}

    for name in possible_names:
        if name.lower() in df_columns_lower:
            return df_columns_lower[name.lower()]

    if required:
        raise ValueError(
            f"Could not find column. Tried: {possible_names}\n"
            f"Available columns: {list(df.columns)}"
        )
    return None


def clean_url(url):
    """Clean a URL by removing tracking parameters and non-standard characters."""
    if pd.isna(url):
        return None

    url = str(url).strip()

    # Remove URLs with tracking parameters (containing ? or =)
    if '?' in url or '=' in url:
        return None

    # Remove fragment identifiers
    if '#' in url:
        url = url.split('#')[0]

    # Check for non-ASCII characters (non-English)
    try:
        url.encode('ascii')
    except UnicodeEncodeError:
        return None

    # Remove URLs with common tracking patterns
    tracking_patterns = [
        'utm_', 'fbclid', 'gclid', 'msclkid', 'mc_', 'ref=',
        '_ga=', 'affiliate', 'campaign', 'source='
    ]
    url_lower = url.lower()
    if any(pattern in url_lower for pattern in tracking_patterns):
        return None

    return url


def clean_url_column(df, column_name):
    """Clean URLs in a DataFrame column."""
    df = df.copy()
    df[column_name] = df[column_name].apply(clean_url)
    df = df[df[column_name].notna()]
    return df


# ----------- DATA PROCESSING FUNCTIONS -----------

def clean_link_dataset(df):
    """Clean and process link dataset according to specified rules."""
    df = df.copy()

    # Filter for Hyperlinks
    if 'Type' in df.columns:
        df = df.sort_values('Type')
        df = df[df['Type'] == 'Hyperlink'].drop('Type', axis=1)

    # Filter for status 200
    if 'Status Code' in df.columns:
        df = df.sort_values('Status Code')
        df = df[df['Status Code'] == 200]
        columns_to_drop = ['Status Code', 'Status'] if 'Status' in df.columns else ['Status Code']
        df = df.drop(columns_to_drop, axis=1)

    # Delete unnecessary columns
    columns_to_drop = [
        'Size (Bytes)', 'Follow', 'Target', 'Rel',
        'Path Type', 'Link Path', 'Link Origin'
    ]
    columns_to_drop = [col for col in columns_to_drop if col in df.columns]
    if columns_to_drop:
        df = df.drop(columns_to_drop, axis=1)

    # Filter for Content and Aside positions
    if 'Link Position' in df.columns:
        df = df.sort_values('Link Position')
        df = df[df['Link Position'].isin(['Content', 'Aside'])]

    # Clean Source URLs
    source_col = 'Source' if 'Source' in df.columns else df.columns[0]
    df = df.sort_values(source_col)

    def is_valid_page(url):
        if pd.isna(url):
            return False
        invalid_patterns = [
            'category/', 'tag/', 'sitemap', 'search', '/home/', 'index'
        ]
        return not any(pattern in str(url).lower() for pattern in invalid_patterns)

    df = df[df[source_col].apply(is_valid_page)]

    # Clean Destination URLs
    dest_col = 'Destination' if 'Destination' in df.columns else df.columns[1]
    df = df.sort_values(dest_col)
    df = df[df[dest_col].apply(is_valid_page)]

    # Process Alt Text
    if 'Alt Text' in df.columns and 'Anchor' in df.columns:
        df = df.sort_values('Alt Text', ascending=False)
        df.loc[df['Alt Text'].notna(), 'Anchor'] = df['Alt Text']
        df = df.drop('Alt Text', axis=1)

    # Remove self-links
    df['links to self'] = np.where(df[source_col] == df[dest_col], 'Match', 'No Match')
    df = df.sort_values('links to self')
    df = df[df['links to self'] != 'Match']
    df = df.drop('links to self', axis=1)

    # Remove Link Position column
    if 'Link Position' in df.columns:
        df = df.drop('Link Position', axis=1)

    # Standardize column names
    if source_col != 'Source' or dest_col != 'Destination':
        df = df.rename(columns={source_col: 'Source', dest_col: 'Destination'})

    if 'Anchor' not in df.columns:
        df['Anchor'] = ''

    final_columns = ['Source', 'Destination', 'Anchor']
    other_columns = [col for col in df.columns if col not in final_columns]
    df = df[final_columns + other_columns]

    return df


def clean_embeddings_data(df):
    """Clean and preprocess embeddings data."""
    df = df.copy()

    # Find embeddings column
    embeddings_col = None
    for col in df.columns:
        if 'embeddings' in col.lower() or 'extract' in col.lower():
            embeddings_col = col
            break

    if not embeddings_col:
        raise ValueError(
            "Could not find embeddings column. "
            "Please ensure your CSV has a column with 'embeddings' or 'extract' in its name."
        )

    # Sort and filter invalid embeddings
    df = df.sort_values(embeddings_col, ascending=False)

    def is_valid_embedding(text):
        if pd.isna(text):
            return False
        invalid_words = ['timeout', 'error', 'null', 'undefined', 'nan']
        if any(word in str(text).lower() for word in invalid_words):
            return False
        text_str = str(text)
        has_numbers = any(c.isdigit() for c in text_str)
        has_separators = ',' in text_str or '.' in text_str
        return has_numbers and has_separators

    df = df[df[embeddings_col].apply(is_valid_embedding)]

    # Filter for status 200 if available
    if 'Status Code' in df.columns:
        df = df[df['Status Code'] == 200]

    # Find URL column
    url_col = None
    potential_url_cols = ['URL', 'Address', 'Url', 'address']
    for col in potential_url_cols:
        if col in df.columns:
            url_col = col
            break

    if not url_col:
        for col in df.columns:
            if col != embeddings_col and col != 'Status Code' and col != 'Status':
                url_col = col
                break

    if not url_col:
        raise ValueError("Could not identify a URL column.")

    # Drop status columns
    cols_to_drop = [col for col in ['Status Code', 'Status'] if col in df.columns]
    if cols_to_drop:
        df = df.drop(cols_to_drop, axis=1)

    # Create cleaned DataFrame
    cleaned_df = pd.DataFrame()
    cleaned_df['URL'] = df[url_col]
    cleaned_df['Embeddings'] = df[embeddings_col]

    # Remove HTTP URLs (only keep HTTPS)
    cleaned_df = cleaned_df[cleaned_df['URL'].str.lower().str.startswith('https://')]

    # Filter out paginated URLs, legal pages, utility pages
    cleaned_df = cleaned_df[~cleaned_df['URL'].apply(should_exclude_url)]

    # Deduplicate URLs based on normalized version (removes protocol and www for comparison)
    cleaned_df['URL_normalized'] = cleaned_df['URL'].apply(normalize_url_for_dedup)
    cleaned_df = cleaned_df.drop_duplicates('URL_normalized', keep='first')
    cleaned_df = cleaned_df.drop(['URL_normalized'], axis=1)

    return cleaned_df


def normalize_url_for_dedup(url):
    """Normalize URL for deduplication purposes."""
    if pd.isna(url):
        return ""
    url = str(url).lower().strip()
    url = url.replace('https://', '').replace('http://', '')
    url = url.replace('www.', '')
    url = url.rstrip('/')
    return url


def should_exclude_url(url):
    """Check if URL should be excluded (pagination, legal pages, utility pages)."""
    if pd.isna(url):
        return True

    url_lower = str(url).lower()

    # Paginated URLs (e.g., /page/2/, /page/5/)
    import re
    if re.search(r'/page/\d+/?', url_lower):
        return True

    # Legal and utility pages to exclude
    exclude_patterns = [
        '/privacy-policy', '/privacy', '/privacypolicy',
        '/terms-of-service', '/terms-and-conditions', '/terms', '/tos',
        '/contact-us', '/contact', '/contactus',
        '/about-us', '/about', '/aboutus',
        '/disclaimer', '/legal',
        '/cookie-policy', '/cookies',
        '/sitemap', '/site-map',
        '/search', '/404', '/error',
        '/login', '/register', '/signup', '/sign-up',
        '/cart', '/checkout', '/account', '/my-account',
        '/wp-admin', '/wp-login', '/admin',
        '/feed', '/rss',
        '/author/', '/tag/', '/category/',
        '/reviews', '/testimonials',
    ]

    for pattern in exclude_patterns:
        if pattern in url_lower:
            return True

    return False


def process_gsc_data(df):
    """Process GSC organic performance data with flexible column name handling."""
    df = df.copy()

    # Find and standardize columns
    url_col = find_column(df, GSC_URL_COLUMNS, required=True)
    clicks_col = find_column(df, GSC_CLICKS_COLUMNS, required=True)
    impressions_col = find_column(df, GSC_IMPRESSIONS_COLUMNS, required=True)
    position_col = find_column(df, GSC_POSITION_COLUMNS, required=True)
    ctr_col = find_column(df, GSC_CTR_COLUMNS, required=True)

    # Create standardized DataFrame
    gsc_df = pd.DataFrame()
    gsc_df['URL'] = df[url_col].str.strip()
    gsc_df['Clicks'] = pd.to_numeric(df[clicks_col], errors='coerce').fillna(0).astype(int)
    gsc_df['Impressions'] = pd.to_numeric(df[impressions_col], errors='coerce').fillna(0).astype(int)
    gsc_df['Avg. Position'] = pd.to_numeric(df[position_col], errors='coerce').round(1)

    # Handle CTR
    ctr_values = df[ctr_col].astype(str)
    ctr_values = ctr_values.str.replace('%', '', regex=False)
    gsc_df['CTR'] = pd.to_numeric(ctr_values, errors='coerce')

    if gsc_df['CTR'].max() <= 1:
        gsc_df['CTR'] = (gsc_df['CTR'] * 100).round(2)
    else:
        gsc_df['CTR'] = gsc_df['CTR'].round(2)

    gsc_df['CTR'] = gsc_df['CTR'].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "0.00%")

    # Remove duplicates
    gsc_df = gsc_df.sort_values('Clicks', ascending=False).drop_duplicates('URL', keep='first')

    return gsc_df


def normalize_url(url):
    """Normalize URL for comparison (remove protocol, lowercase, strip trailing slash)."""
    if pd.isna(url):
        return ""
    url = str(url).lower().strip()
    # Remove protocol (http:// or https://)
    url = url.replace('https://', '').replace('http://', '')
    # Remove www. prefix
    url = url.replace('www.', '')
    # Strip trailing slash
    url = url.rstrip('/')
    return url


def find_related_pages(df, top_n=5, min_similarity=0.0, progress_callback=None):
    """Find top N related pages based on cosine similarity."""
    related_pages = {}
    embeddings = np.stack(df['Embeddings'].values)
    urls = df['URL'].values

    # Create normalized URL list for comparison
    normalized_urls = [normalize_url(u) for u in urls]

    cosine_similarities = cosine_similarity(embeddings)

    for idx, url in enumerate(urls):
        if progress_callback and idx % 10 == 0:
            progress_callback(idx / len(urls))

        similar_indices = cosine_similarities[idx].argsort()[::-1]
        current_url_normalized = normalized_urls[idx]

        related_with_scores = []
        for sim_idx in similar_indices:
            # Skip if same URL (using normalized comparison)
            if normalized_urls[sim_idx] == current_url_normalized:
                continue

            score = cosine_similarities[idx][sim_idx]
            if score >= min_similarity:
                related_with_scores.append((urls[sim_idx], round(score, 4)))
                if len(related_with_scores) >= top_n:
                    break

        related_pages[url] = related_with_scores

    if progress_callback:
        progress_callback(1.0)

    return related_pages


def create_excel_output(final_df, top_n):
    """Create Excel file with formatting and return as bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Internal Link Opportunities'

    headers = list(final_df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, row in enumerate(final_df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center')

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")

    base_cols = 5
    link_status_cols = [base_cols + (i * 3) + 3 for i in range(top_n)]

    for col_idx in link_status_cols:
        for row_idx in range(2, len(final_df) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value == "Exists":
                cell.fill = green_fill
            elif cell.value == "Not Found":
                cell.fill = red_fill

    base_widths = [50, 10, 12, 14, 10]
    for col_idx, width in enumerate(base_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for i in range(top_n):
        col_base = base_cols + (i * 3) + 1
        ws.column_dimensions[get_column_letter(col_base)].width = 50
        ws.column_dimensions[get_column_letter(col_base + 1)].width = 12
        ws.column_dimensions[get_column_letter(col_base + 2)].width = 18

    ws.freeze_panes = 'B2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ----------- STREAMLIT UI -----------

def main():
    # Header
    st.title("🔗 Internal Link Finder")
    st.markdown("Discover internal linking opportunities using vector embeddings and GSC metrics")

    # Sidebar settings
    with st.sidebar:
        st.header("⚙️ Settings")

        top_n = st.slider(
            "Related pages per URL",
            min_value=1,
            max_value=20,
            value=5,
            help="Number of semantically similar pages to find for each URL"
        )

        min_similarity = st.slider(
            "Minimum similarity threshold",
            min_value=0.0,
            max_value=1.0,
            value=0.0,
            step=0.05,
            help="Only show pages with similarity above this threshold. 0.7+ recommended for quality matches."
        )

        clean_urls = st.checkbox(
            "Clean URLs",
            value=True,
            help="Remove URLs with tracking parameters (?, =) and non-English characters"
        )

        st.divider()

        st.header("📚 About")
        st.markdown("""
        This tool analyzes your website's internal link structure using:

        1. **Screaming Frog Inlinks** - Existing link structure
        2. **Embeddings Report** - Semantic similarity via OpenAI
        3. **GSC Performance** - Traffic metrics for prioritization

        Based on [Everett Sizemore's methodology](https://moz.com/blog/internal-linking-opportunities-with-vector-embeddings).
        """)

    # File upload section
    st.header("📁 Upload Your Files")

    col1, col2, col3 = st.columns(3)

    with col1:
        st.subheader("1. Inlinks Report")
        inlinks_file = st.file_uploader(
            "Screaming Frog All Inlinks CSV",
            type=['csv'],
            key="inlinks",
            help="Bulk Export > Links > All Inlinks"
        )

    with col2:
        st.subheader("2. Embeddings Report")
        embeddings_file = st.file_uploader(
            "Screaming Frog Embeddings CSV",
            type=['csv'],
            key="embeddings",
            help="Bulk Export > Custom > Custom JavaScript"
        )

    with col3:
        st.subheader("3. GSC Performance")
        gsc_file = st.file_uploader(
            "GSC Organic Performance CSV",
            type=['csv'],
            key="gsc",
            help="Export from Google Search Console"
        )

    # Process button
    if st.button("🚀 Find Link Opportunities", type="primary", use_container_width=True):
        if not all([inlinks_file, embeddings_file, gsc_file]):
            st.error("Please upload all three required files.")
            return

        try:
            progress_bar = st.progress(0, text="Starting analysis...")
            status_text = st.empty()

            # Step 1: Process inlinks
            status_text.text("Processing inlinks data...")
            progress_bar.progress(10, text="Processing inlinks...")

            df_links = pd.read_csv(inlinks_file)
            df_cleaned_links = clean_link_dataset(df_links)

            if clean_urls:
                df_cleaned_links = clean_url_column(df_cleaned_links, 'Source')
                df_cleaned_links = clean_url_column(df_cleaned_links, 'Destination')

            # Step 2: Process embeddings
            status_text.text("Processing embeddings data...")
            progress_bar.progress(30, text="Processing embeddings...")

            df_embeddings_raw = pd.read_csv(embeddings_file)
            df_embeddings = clean_embeddings_data(df_embeddings_raw)

            if clean_urls:
                df_embeddings = clean_url_column(df_embeddings, 'URL')

            df_embeddings['Embeddings'] = df_embeddings['Embeddings'].apply(
                lambda x: np.array([float(i) for i in str(x).strip('[]').replace("'", "").split(',')])
            )

            # Step 3: Process GSC data
            status_text.text("Processing GSC data...")
            progress_bar.progress(50, text="Processing GSC data...")

            df_gsc_raw = pd.read_csv(gsc_file)
            df_gsc = process_gsc_data(df_gsc_raw)

            if clean_urls:
                df_gsc = clean_url_column(df_gsc, 'URL')

            # Step 4: Find related pages
            status_text.text("Calculating semantic similarity...")
            progress_bar.progress(60, text="Analyzing relationships...")

            def update_progress(pct):
                progress_bar.progress(int(60 + pct * 30), text=f"Finding related pages... {int(pct*100)}%")

            related_pages = find_related_pages(
                df_embeddings,
                top_n=top_n,
                min_similarity=min_similarity,
                progress_callback=update_progress
            )

            # Step 5: Build output
            status_text.text("Building results...")
            progress_bar.progress(90, text="Creating output...")

            output_data = []

            for url, related_with_scores in related_pages.items():
                padded_related = related_with_scores + [(None, None)] * (top_n - len(related_with_scores))

                gsc_row = df_gsc[df_gsc['URL'] == url]

                if not gsc_row.empty:
                    clicks = gsc_row['Clicks'].values[0]
                    impressions = gsc_row['Impressions'].values[0]
                    avg_position = gsc_row['Avg. Position'].values[0]
                    ctr = gsc_row['CTR'].values[0]
                else:
                    clicks = 0
                    impressions = 0
                    avg_position = None
                    ctr = "0.00%"

                row = {
                    'URL': url,
                    'Clicks': clicks,
                    'Impressions': impressions,
                    'Avg. Position': avg_position if pd.notna(avg_position) else '',
                    'CTR': ctr
                }

                for i, (related_url, similarity_score) in enumerate(padded_related, 1):
                    row[f'Related URL {i}'] = related_url
                    row[f'Similarity {i}'] = f"{similarity_score:.2%}" if similarity_score is not None else ""
                    if related_url is not None:
                        exists_status = (
                            "Exists"
                            if related_url in df_cleaned_links[
                                df_cleaned_links['Destination'] == url
                            ]['Source'].values
                            else "Not Found"
                        )
                        row[f'URL {i} links to Target?'] = exists_status
                    else:
                        row[f'URL {i} links to Target?'] = ""

                output_data.append(row)

            column_order = ['URL', 'Clicks', 'Impressions', 'Avg. Position', 'CTR']
            for i in range(1, top_n + 1):
                column_order.extend([f'Related URL {i}', f'Similarity {i}', f'URL {i} links to Target?'])

            final_df = pd.DataFrame(output_data)
            final_df = final_df[column_order]
            final_df = final_df.sort_values('Clicks', ascending=False).reset_index(drop=True)

            progress_bar.progress(100, text="Complete!")
            status_text.empty()

            # Results
            st.success("✅ Analysis complete!")

            # Summary metrics
            col1, col2, col3, col4 = st.columns(4)

            total_opportunities = sum(
                (final_df[f'URL {i} links to Target?'] == 'Not Found').sum()
                for i in range(1, top_n + 1)
            )

            with col1:
                st.metric("URLs Analyzed", len(final_df))
            with col2:
                st.metric("URLs with Traffic", len(final_df[final_df['Clicks'] > 0]))
            with col3:
                st.metric("Link Opportunities", total_opportunities)
            with col4:
                existing_links = sum(
                    (final_df[f'URL {i} links to Target?'] == 'Exists').sum()
                    for i in range(1, top_n + 1)
                )
                st.metric("Existing Links", existing_links)

            # Download buttons
            st.header("📥 Download Results")

            col1, col2 = st.columns(2)

            with col1:
                excel_data = create_excel_output(final_df, top_n)
                st.download_button(
                    label="📊 Download Excel (with formatting)",
                    data=excel_data,
                    file_name="internal_link_opportunities.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            with col2:
                csv_data = final_df.to_csv(index=False)
                st.download_button(
                    label="📄 Download CSV",
                    data=csv_data,
                    file_name="internal_link_opportunities.csv",
                    mime="text/csv",
                    use_container_width=True
                )

            # Preview
            st.header("📋 Results Preview")

            preview_cols = ['URL', 'Clicks', 'Impressions', 'Related URL 1', 'Similarity 1', 'URL 1 links to Target?']
            st.dataframe(final_df[preview_cols].head(50), use_container_width=True, hide_index=True)

            with st.expander("View Full Results"):
                st.dataframe(final_df, use_container_width=True, hide_index=True)

        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
            st.exception(e)

    # Instructions
    with st.expander("📖 How to Get the Required Files"):
        st.markdown("""
        ### 1. Screaming Frog Inlinks Report
        1. Open Screaming Frog and crawl your website
        2. Go to **Bulk Export > Links > All Inlinks**
        3. Save the CSV file

        ### 2. Screaming Frog Embeddings Report
        1. Go to **Configuration > API Access > AI** and add your OpenAI API key
        2. Go to **Configuration > Custom > Custom JavaScript**
        3. Click **Add from Library** > Select **(ChatGPT) Extract embeddings from page content**
        4. Run your crawl
        5. Go to **Bulk Export > Custom > Custom JavaScript**

        ### 3. GSC Organic Performance Report
        1. Go to [Google Search Console](https://search.google.com/search-console)
        2. Navigate to **Performance > Search Results**
        3. Click on **Pages** tab
        4. Export as CSV
        """)


if __name__ == "__main__":
    main()
