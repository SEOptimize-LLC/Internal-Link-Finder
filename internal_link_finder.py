#!/usr/bin/env python3
"""
Internal Link Opportunities Finder with GSC Integration

Based on Everett Sizemore's tutorial:
"How I Found Internal Linking Opportunities With Vector Embeddings"
https://moz.com/blog/internal-linking-opportunities-with-vector-embeddings

Original script by Britney Muller
Enhanced with GSC integration by SEOptimize-LLC

This script performs:
1. Raw all_inlinks cleaning from Screaming Frog
2. Raw vector embeddings preprocessing and cleaning
3. GSC organic performance data integration
4. Finds contextually related pages using vector embeddings
5. Analyzes + identifies top link opportunities between related pages
"""

import pandas as pd
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
import os
import sys

# Check if running in Google Colab
try:
    from google.colab import files
    IN_COLAB = True
except ImportError:
    IN_COLAB = False


# ----------- COLUMN NAME MAPPINGS -----------

# GSC column name variations mapping
GSC_URL_COLUMNS = [
    'landing page', 'landing pages', 'url', 'urls', 'page', 'pages',
    'address', 'landing_page', 'landing_pages'
]

GSC_CLICKS_COLUMNS = [
    'clicks', 'click', 'total clicks', 'total_clicks'
]

GSC_IMPRESSIONS_COLUMNS = [
    'impressions', 'impression', 'total impressions', 'total_impressions', 'impr'
]

GSC_POSITION_COLUMNS = [
    'average position', 'avg. position', 'avg position', 'avg. pos', 'avg pos',
    'position', 'pos', 'average_position', 'avg_position', 'avg_pos',
    'mean position', 'mean_position'
]

GSC_CTR_COLUMNS = [
    'ctr', 'url ctr', 'average ctr', 'avg. ctr', 'avg ctr',
    'click through rate', 'click_through_rate', 'url_ctr', 'avg_ctr'
]


# ----------- HELPER FUNCTIONS -----------

def find_column(df, possible_names, required=True):
    """
    Find a column in the dataframe that matches any of the possible names.
    Case-insensitive matching.

    Parameters:
    df: DataFrame to search
    possible_names: List of possible column names (lowercase)
    required: If True, raise error when not found

    Returns:
    Column name if found, None otherwise
    """
    df_columns_lower = {col.lower().strip(): col for col in df.columns}

    for name in possible_names:
        if name.lower() in df_columns_lower:
            return df_columns_lower[name.lower()]

    if required:
        raise ValueError(
            f"Could not find column. Tried: {possible_names}\n"
            f"Available columns: {list(df.columns)}"
        )
    return None


def upload_file(prompt_message):
    """Upload a file (Colab) or prompt for file path (local)."""
    if IN_COLAB:
        print(prompt_message)
        uploaded = files.upload()
        if not uploaded:
            raise ValueError("No file was uploaded.")
        file_name = list(uploaded.keys())[0]
        print(f"Uploaded: {file_name}")
        return file_name
    else:
        print(prompt_message)
        file_path = input("Enter the file path: ").strip()
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        return file_path


def download_file(file_path):
    """Download file (Colab) or print location (local)."""
    if IN_COLAB:
        files.download(file_path)
    else:
        print(f"File saved: {os.path.abspath(file_path)}")


def clean_url(url):
    """
    Clean a URL by removing tracking parameters and non-standard characters.

    Returns:
    str: Cleaned URL or None if URL should be excluded
    """
    if pd.isna(url):
        return None

    url = str(url).strip()

    # Remove URLs with tracking parameters (containing ? or =)
    if '?' in url or '=' in url:
        return None

    # Remove fragment identifiers
    if '#' in url:
        url = url.split('#')[0]

    # Check for non-ASCII characters (non-English)
    try:
        url.encode('ascii')
    except UnicodeEncodeError:
        return None

    # Remove URLs with common tracking patterns
    tracking_patterns = [
        'utm_', 'fbclid', 'gclid', 'msclkid', 'mc_', 'ref=',
        '_ga=', 'affiliate', 'campaign', 'source='
    ]
    url_lower = url.lower()
    if any(pattern in url_lower for pattern in tracking_patterns):
        return None

    return url


def clean_url_column(df, column_name, description=""):
    """
    Clean URLs in a DataFrame column.

    Parameters:
    df: DataFrame
    column_name: Name of the column containing URLs
    description: Description for logging

    Returns:
    DataFrame with cleaned URLs (invalid URLs removed)
    """
    original_count = len(df)

    # Apply URL cleaning
    df = df.copy()
    df[column_name] = df[column_name].apply(clean_url)

    # Remove rows where URL became None
    df = df[df[column_name].notna()]

    removed_count = original_count - len(df)
    if removed_count > 0:
        print(f"  {description}: Removed {removed_count} URLs with tracking params or non-standard characters")

    return df


# ----------- PART 1: LINK DATASET CLEANING -----------

def clean_link_dataset(df):
    """
    Clean and process link dataset according to specified rules.

    Parameters:
    df (pandas.DataFrame): Input DataFrame containing link data

    Returns:
    pandas.DataFrame: Cleaned DataFrame with Source, Destination, and Anchor columns
    """
    df = df.copy()
    print("Initial shape:", df.shape)

    # 1. Sort by Type and filter for Hyperlinks
    print("\nSorting by Type and filtering for Hyperlinks...")
    if 'Type' in df.columns:
        df = df.sort_values('Type')
        df = df[df['Type'] == 'Hyperlink'].drop('Type', axis=1)
        print("Shape after Type filtering:", df.shape)

    # 2. Sort by Status Code and filter for 200
    print("\nSorting by Status Code and filtering for 200 status...")
    if 'Status Code' in df.columns:
        df = df.sort_values('Status Code')
        df = df[df['Status Code'] == 200]
        columns_to_drop = ['Status Code', 'Status'] if 'Status' in df.columns else ['Status Code']
        df = df.drop(columns_to_drop, axis=1)
        print("Shape after Status filtering:", df.shape)

    # 3. Delete specified columns if they exist
    columns_to_drop = [
        'Size (Bytes)', 'Follow', 'Target', 'Rel',
        'Path Type', 'Link Path', 'Link Origin'
    ]
    columns_to_drop = [col for col in columns_to_drop if col in df.columns]

    if columns_to_drop:
        print("\nRemoving unnecessary columns...")
        df = df.drop(columns_to_drop, axis=1)
    print("Remaining columns:", df.columns.tolist())

    # 4. Sort by Link Position
    if 'Link Position' in df.columns:
        print("\nSorting by Link Position...")
        df = df.sort_values('Link Position')

        # 5. Filter for Content and Aside in Link Position
        print("\nFiltering for Content and Aside positions...")
        df = df[df['Link Position'].isin(['Content', 'Aside'])]
        print("Shape after Link Position filtering:", df.shape)

    # 6. Sort by Source and clean Source URLs
    print("\nSorting and cleaning Source URLs...")
    source_col = 'Source' if 'Source' in df.columns else df.columns[0]
    df = df.sort_values(source_col)

    def is_valid_page(url):
        if pd.isna(url):
            return False
        # Customize these patterns as needed for your site
        invalid_patterns = [
            'category/', 'tag/', 'sitemap', 'search', '/home/', 'index'
        ]
        return not any(pattern in str(url).lower() for pattern in invalid_patterns)

    df = df[df[source_col].apply(is_valid_page)]
    print(f"Shape after {source_col} URL cleaning:", df.shape)

    # 7. Sort by Destination and clean Destination URLs
    print("\nSorting and cleaning Destination URLs...")
    dest_col = 'Destination' if 'Destination' in df.columns else df.columns[1]
    df = df.sort_values(dest_col)
    df = df[df[dest_col].apply(is_valid_page)]
    print(f"Shape after {dest_col} URL cleaning:", df.shape)

    # 8. Sort by Alt Text (Z to A) and process Alt Text if it exists
    if 'Alt Text' in df.columns and 'Anchor' in df.columns:
        print("\nSorting by Alt Text and processing...")
        df = df.sort_values('Alt Text', ascending=False)
        df.loc[df['Alt Text'].notna(), 'Anchor'] = df['Alt Text']
        df = df.drop('Alt Text', axis=1)

    # 9. Handle self-linking URLs
    print("\nProcessing self-linking URLs...")
    df['links to self'] = np.where(df[source_col] == df[dest_col], 'Match', 'No Match')
    df = df.sort_values('links to self')
    df = df[df['links to self'] != 'Match']
    df = df.drop('links to self', axis=1)
    print("Shape after removing self-links:", df.shape)

    # Remove Link Position column if it exists
    if 'Link Position' in df.columns:
        df = df.drop('Link Position', axis=1)

    # Ensure we have the expected columns
    if source_col != 'Source' or dest_col != 'Destination':
        df = df.rename(columns={source_col: 'Source', dest_col: 'Destination'})

    if 'Anchor' not in df.columns:
        df['Anchor'] = ''

    # Final column ordering
    final_columns = ['Source', 'Destination', 'Anchor']
    other_columns = [col for col in df.columns if col not in final_columns]
    df = df[final_columns + other_columns]

    return df


# ----------- PART 2: EMBEDDINGS PREPROCESSING -----------

def clean_embeddings_data(df):
    """
    Clean and preprocess embeddings data according to specified rules.

    Parameters:
    df (pandas.DataFrame): Input DataFrame containing embeddings data

    Returns:
    pandas.DataFrame: Cleaned DataFrame with URL and Embeddings columns
    """
    print("\n--- CLEANING EMBEDDINGS DATA ---")
    df = df.copy()
    print("Initial shape:", df.shape)

    # Find the embeddings column
    embeddings_col = None
    for col in df.columns:
        if 'embeddings' in col.lower() or 'extract' in col.lower():
            embeddings_col = col
            break

    if not embeddings_col:
        raise ValueError(
            "Could not find embeddings column. "
            "Please ensure your CSV has a column with 'embeddings' or 'extract' in its name."
        )

    # Sort the embeddings column from Z to A
    print(f"\nSorting {embeddings_col} column from Z to A...")
    df = df.sort_values(embeddings_col, ascending=False)

    # Delete rows with invalid embeddings
    print("\nRemoving rows with invalid embeddings data...")

    def is_valid_embedding(text):
        if pd.isna(text):
            return False
        invalid_words = ['timeout', 'error', 'null', 'undefined', 'nan']
        if any(word in str(text).lower() for word in invalid_words):
            return False
        text_str = str(text)
        has_numbers = any(c.isdigit() for c in text_str)
        has_separators = ',' in text_str or '.' in text_str
        return has_numbers and has_separators

    df = df[df[embeddings_col].apply(is_valid_embedding)]
    print("Shape after removing invalid embeddings:", df.shape)

    # Filter for status code 200 if available
    if 'Status Code' in df.columns:
        print("\nFiltering for status code 200...")
        df = df[df['Status Code'] == 200]
        print("Shape after status code filtering:", df.shape)

    # Find the URL column
    url_col = None
    potential_url_cols = ['URL', 'Address', 'Url', 'address']

    for col in potential_url_cols:
        if col in df.columns:
            url_col = col
            break

    if not url_col:
        for col in df.columns:
            if col != embeddings_col and col != 'Status Code' and col != 'Status':
                url_col = col
                break

    if not url_col:
        raise ValueError("Could not identify a URL column.")

    # Drop status columns
    cols_to_drop = [col for col in ['Status Code', 'Status'] if col in df.columns]
    if cols_to_drop:
        df = df.drop(cols_to_drop, axis=1)

    # Create cleaned DataFrame
    cleaned_df = pd.DataFrame()
    cleaned_df['URL'] = df[url_col]
    cleaned_df['Embeddings'] = df[embeddings_col]

    print("\nFinal embeddings data shape:", cleaned_df.shape)
    print("Final columns:", cleaned_df.columns.tolist())

    return cleaned_df


# ----------- PART 3: GSC DATA PROCESSING -----------

def process_gsc_data(df):
    """
    Process GSC organic performance data with flexible column name handling.

    Parameters:
    df (pandas.DataFrame): Input DataFrame containing GSC data

    Returns:
    pandas.DataFrame: Processed DataFrame with standardized column names
    """
    print("\n--- PROCESSING GSC DATA ---")
    df = df.copy()
    print("Initial shape:", df.shape)
    print("Available columns:", df.columns.tolist())

    # Find and standardize columns
    url_col = find_column(df, GSC_URL_COLUMNS, required=True)
    clicks_col = find_column(df, GSC_CLICKS_COLUMNS, required=True)
    impressions_col = find_column(df, GSC_IMPRESSIONS_COLUMNS, required=True)
    position_col = find_column(df, GSC_POSITION_COLUMNS, required=True)
    ctr_col = find_column(df, GSC_CTR_COLUMNS, required=True)

    print(f"\nIdentified columns:")
    print(f"  URL column: {url_col}")
    print(f"  Clicks column: {clicks_col}")
    print(f"  Impressions column: {impressions_col}")
    print(f"  Position column: {position_col}")
    print(f"  CTR column: {ctr_col}")

    # Create standardized DataFrame
    gsc_df = pd.DataFrame()
    gsc_df['URL'] = df[url_col].str.strip()
    gsc_df['Clicks'] = pd.to_numeric(df[clicks_col], errors='coerce').fillna(0).astype(int)
    gsc_df['Impressions'] = pd.to_numeric(df[impressions_col], errors='coerce').fillna(0).astype(int)

    # Handle position (round to 1 decimal)
    gsc_df['Avg. Position'] = pd.to_numeric(df[position_col], errors='coerce').round(1)

    # Handle CTR (convert percentage strings if needed)
    ctr_values = df[ctr_col].astype(str)
    # Remove % sign if present and convert
    ctr_values = ctr_values.str.replace('%', '', regex=False)
    gsc_df['CTR'] = pd.to_numeric(ctr_values, errors='coerce')

    # If CTR values are > 1, they're likely already percentages, otherwise multiply by 100
    if gsc_df['CTR'].max() <= 1:
        gsc_df['CTR'] = (gsc_df['CTR'] * 100).round(2)
    else:
        gsc_df['CTR'] = gsc_df['CTR'].round(2)

    # Format CTR as string with % sign
    gsc_df['CTR'] = gsc_df['CTR'].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "0.00%")

    # Remove duplicates, keeping the row with highest clicks
    gsc_df = gsc_df.sort_values('Clicks', ascending=False).drop_duplicates('URL', keep='first')

    print(f"\nProcessed GSC data shape: {gsc_df.shape}")

    return gsc_df


# ----------- PART 4: URL RELATIONSHIP ANALYSIS -----------

# Default minimum similarity threshold (0.0 = no filtering, 0.7+ recommended for quality)
DEFAULT_MIN_SIMILARITY = 0.0


def find_related_pages(df, top_n=5, min_similarity=DEFAULT_MIN_SIMILARITY):
    """
    Find top N related pages based on cosine similarity.

    Parameters:
    df: DataFrame with URL and Embeddings columns
    top_n: Number of related pages to find
    min_similarity: Minimum cosine similarity threshold (0.0-1.0)
                   Pages with similarity below this are excluded.
                   Recommended: 0.7+ for high-quality matches

    Returns:
    dict: {url: [(related_url, similarity_score), ...]}
    """
    print(f"Finding top {top_n} related pages for each URL...")
    if min_similarity > 0:
        print(f"Minimum similarity threshold: {min_similarity:.2f}")

    related_pages = {}
    embeddings = np.stack(df['Embeddings'].values)
    urls = df['URL'].values

    # Calculate cosine similarity matrix
    cosine_similarities = cosine_similarity(embeddings)

    # For each URL, find the most similar URLs
    for idx, url in enumerate(urls):
        # Get indices sorted by similarity (descending)
        similar_indices = cosine_similarities[idx].argsort()[::-1]

        # Filter out self and apply threshold
        related_with_scores = []
        for sim_idx in similar_indices:
            if urls[sim_idx] != url:
                score = cosine_similarities[idx][sim_idx]
                if score >= min_similarity:
                    related_with_scores.append((urls[sim_idx], round(score, 4)))
                    if len(related_with_scores) >= top_n:
                        break

        related_pages[url] = related_with_scores

    # Report statistics
    total_relations = sum(len(v) for v in related_pages.values())
    avg_relations = total_relations / len(related_pages) if related_pages else 0
    print(f"Average related pages per URL: {avg_relations:.1f}")

    if min_similarity > 0:
        urls_with_fewer = sum(1 for v in related_pages.values() if len(v) < top_n)
        if urls_with_fewer > 0:
            print(f"Note: {urls_with_fewer} URLs have fewer than {top_n} matches above threshold")

    return related_pages


# ----------- PART 5: MAIN PROCESS -----------

def main(min_similarity=DEFAULT_MIN_SIMILARITY, top_n=5, clean_urls=True):
    """
    Main process for internal link opportunity analysis.

    Parameters:
    min_similarity: Minimum cosine similarity threshold (0.0-1.0)
    top_n: Number of related pages to find per URL
    clean_urls: If True, remove URLs with tracking parameters and non-English characters
    """

    print("=" * 80)
    print("Internal Linking Analysis Tool with GSC Integration")
    print("Combining link cleaning, embeddings analysis, GSC metrics, and relationship discovery")
    print("=" * 80)
    if min_similarity > 0:
        print(f"Similarity threshold: {min_similarity:.0%}")
    print(f"Related pages per URL: {top_n}")
    if clean_urls:
        print("URL cleaning: Enabled (removing tracking parameters and non-standard characters)")

    # Step 1: Upload and clean link data
    print("\n--- STEP 1: CLEANING ALL_INLINKS DATA ---")
    link_export_file = upload_file(
        "Please upload your Screaming Frog 'all_inlinks' export CSV file..."
    )

    df_links = pd.read_csv(link_export_file)
    print("\nLink export file information:")
    print(f"Columns: {df_links.columns.tolist()}")
    print(f"Shape: {df_links.shape}")

    print("\nCleaning link dataset...")
    df_cleaned_links = clean_link_dataset(df_links)

    # Apply URL cleaning if enabled
    if clean_urls:
        print("\nApplying URL cleaning to link data...")
        df_cleaned_links = clean_url_column(df_cleaned_links, 'Source', 'Source URLs')
        df_cleaned_links = clean_url_column(df_cleaned_links, 'Destination', 'Destination URLs')

    # Save cleaned links for reference
    cleaned_links_file = 'cleaned_links.csv'
    df_cleaned_links.to_csv(cleaned_links_file, index=False)
    print(f"\nCleaned link data saved to {cleaned_links_file}")

    # Step 2: Upload and process embeddings data
    print("\n--- STEP 2: EMBEDDINGS PROCESSING ---")
    embeddings_file = upload_file(
        "Please upload your Screaming Frog embeddings CSV export..."
    )

    df_embeddings_raw = pd.read_csv(embeddings_file)
    print(f"Embeddings file columns: {df_embeddings_raw.columns.tolist()}")

    df_embeddings = clean_embeddings_data(df_embeddings_raw)

    # Apply URL cleaning if enabled
    if clean_urls:
        print("\nApplying URL cleaning to embeddings data...")
        df_embeddings = clean_url_column(df_embeddings, 'URL', 'Embeddings URLs')

    cleaned_embeddings_file = 'cleaned_embeddings.csv'
    df_embeddings.to_csv(cleaned_embeddings_file, index=False)
    print(f"\nCleaned embeddings data saved to {cleaned_embeddings_file}")

    # Step 3: Upload and process GSC data
    print("\n--- STEP 3: GSC ORGANIC PERFORMANCE DATA ---")
    gsc_file = upload_file(
        "Please upload your GSC organic performance report CSV file...\n"
        "(Should contain: Landing Pages/URLs, Clicks, Impressions, Avg. Position, CTR)"
    )

    df_gsc_raw = pd.read_csv(gsc_file)
    print(f"GSC file columns: {df_gsc_raw.columns.tolist()}")

    df_gsc = process_gsc_data(df_gsc_raw)

    # Apply URL cleaning if enabled
    if clean_urls:
        print("\nApplying URL cleaning to GSC data...")
        df_gsc = clean_url_column(df_gsc, 'URL', 'GSC URLs')

    # Step 4: Convert embeddings and analyze relationships
    print("\n--- STEP 4: ANALYZING RELATIONSHIPS BETWEEN URLS ---")
    print("Converting embeddings to arrays...")
    df_embeddings['Embeddings'] = df_embeddings['Embeddings'].apply(
        lambda x: np.array([float(i) for i in str(x).strip('[]').replace("'", "").split(',')])
    )

    # Find related pages
    related_pages = find_related_pages(df_embeddings, top_n=top_n, min_similarity=min_similarity)

    # Step 5: Build output DataFrame
    print("\nCreating internal linking opportunities DataFrame...")
    output_data = []

    for url, related_with_scores in related_pages.items():
        # Pad related URLs to always have top_n items (with None for both URL and score)
        padded_related = related_with_scores + [(None, None)] * (top_n - len(related_with_scores))

        # Get GSC metrics for this URL
        gsc_row = df_gsc[df_gsc['URL'] == url]

        if not gsc_row.empty:
            clicks = gsc_row['Clicks'].values[0]
            impressions = gsc_row['Impressions'].values[0]
            avg_position = gsc_row['Avg. Position'].values[0]
            ctr = gsc_row['CTR'].values[0]
        else:
            clicks = 0
            impressions = 0
            avg_position = None
            ctr = "0.00%"

        # Create row - NOTE: "Links to Target URL" column is removed as requested
        row = {
            'URL': url,
            'Clicks': clicks,
            'Impressions': impressions,
            'Avg. Position': avg_position if pd.notna(avg_position) else '',
            'CTR': ctr
        }

        # Add related URLs with similarity scores and link status checks
        for i, (related_url, similarity_score) in enumerate(padded_related, 1):
            row[f'Related URL {i}'] = related_url
            row[f'Similarity {i}'] = f"{similarity_score:.2%}" if similarity_score is not None else ""
            if related_url is not None:
                exists_status = (
                    "Exists"
                    if related_url in df_cleaned_links[
                        df_cleaned_links['Destination'] == url
                    ]['Source'].values
                    else "Not Found"
                )
                row[f'URL {i} links to Target?'] = exists_status
            else:
                row[f'URL {i} links to Target?'] = ""

        output_data.append(row)

    # Create final DataFrame with columns in correct order
    column_order = ['URL', 'Clicks', 'Impressions', 'Avg. Position', 'CTR']
    for i in range(1, top_n + 1):
        column_order.extend([f'Related URL {i}', f'Similarity {i}', f'URL {i} links to Target?'])

    final_df = pd.DataFrame(output_data)
    final_df = final_df[column_order]

    # Sort by Clicks descending (most important pages first)
    final_df = final_df.sort_values('Clicks', ascending=False).reset_index(drop=True)

    # Step 6: Save to Excel with formatting
    output_file_name = 'internal_link_opportunities.xlsx'
    print(f"\nCreating Excel file with formatting: {output_file_name}...")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Internal Link Opportunities'

    # Write header
    headers = list(final_df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data
    for row_idx, row in enumerate(final_df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center')

    # Define fills
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Style header row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")

    # Apply conditional formatting to "URL X links to Target?" columns
    # Column layout: URL, Clicks, Impressions, Avg.Position, CTR (5 cols), then for each 1-N:
    #   Related URL X, Similarity X, URL X links to Target?
    # Status columns start at position 8 (5 base cols + 3rd col of each triplet)
    base_cols = 5
    link_status_cols = [base_cols + (i * 3) + 3 for i in range(top_n)]

    for col_idx in link_status_cols:
        for row_idx in range(2, len(final_df) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value == "Exists":
                cell.fill = green_fill
            elif cell.value == "Not Found":
                cell.fill = red_fill

    # Adjust column widths
    from openpyxl.utils import get_column_letter

    # Set widths for base columns (URL and GSC metrics)
    base_widths = [50, 10, 12, 14, 10]  # URL, Clicks, Impressions, Avg.Position, CTR
    for col_idx, width in enumerate(base_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Set widths for related URL, similarity, and status columns
    for i in range(top_n):
        col_base = base_cols + (i * 3) + 1
        ws.column_dimensions[get_column_letter(col_base)].width = 50      # Related URL
        ws.column_dimensions[get_column_letter(col_base + 1)].width = 12  # Similarity
        ws.column_dimensions[get_column_letter(col_base + 2)].width = 18  # Status

    # Freeze top row and first column
    ws.freeze_panes = 'B2'

    # Save workbook
    wb.save(output_file_name)
    print(f"Excel file saved: {output_file_name}")

    # Also save CSV version
    csv_file_name = 'internal_link_opportunities.csv'
    final_df.to_csv(csv_file_name, index=False)
    print(f"CSV file saved: {csv_file_name}")

    # Download files
    print("\nDownloading output files...")
    download_file(output_file_name)
    download_file(csv_file_name)

    # Print summary
    print("\n" + "=" * 80)
    print("PROCESS COMPLETED SUCCESSFULLY!")
    print("=" * 80)
    print(f"\nOutput files:")
    print(f"  1. {output_file_name} - Excel file with color formatting")
    print(f"  2. {csv_file_name} - CSV file for compatibility")
    print(f"\nSummary:")
    print(f"  - Total URLs analyzed: {len(final_df)}")
    print(f"  - URLs with GSC data: {len(final_df[final_df['Clicks'] > 0])}")

    # Count opportunities
    total_opportunities = 0
    for i in range(1, top_n + 1):
        col = f'URL {i} links to Target?'
        total_opportunities += (final_df[col] == 'Not Found').sum()

    print(f"  - Total linking opportunities found: {total_opportunities}")
    print("\nThe output shows:")
    print("  * Your target URLs with GSC performance metrics")
    print("  * The top 5 contextually similar pages for each target")
    print("  * Cosine similarity scores (higher = more relevant)")
    print("  * Whether each related page links to the target (color-coded)")
    print("    - Green: Link exists")
    print("    - Red: Link opportunity (Not Found)")
    print("\nTip: Higher similarity scores (80%+) indicate stronger semantic")
    print("     relationships and better internal linking opportunities.")

    return final_df


def parse_args():
    """Parse command line arguments."""
    import argparse
    parser = argparse.ArgumentParser(
        description='Internal Link Opportunities Finder with GSC Integration',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python internal_link_finder.py                  # Run with default settings
  python internal_link_finder.py --threshold 0.7  # Only show matches with 70%+ similarity
  python internal_link_finder.py --top-n 10       # Show top 10 related pages instead of 5
  python internal_link_finder.py --no-clean       # Don't clean URLs (keep tracking params)
        """
    )
    parser.add_argument(
        '--threshold', '-t',
        type=float,
        default=DEFAULT_MIN_SIMILARITY,
        help='Minimum similarity threshold (0.0-1.0). Default: 0.0 (no filtering). '
             'Recommended: 0.7+ for high-quality matches only.'
    )
    parser.add_argument(
        '--top-n', '-n',
        type=int,
        default=5,
        help='Number of related pages to find per URL. Default: 5'
    )
    parser.add_argument(
        '--no-clean',
        action='store_true',
        default=False,
        help='Disable URL cleaning (keep URLs with tracking parameters and non-English characters)'
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    main(min_similarity=args.threshold, top_n=args.top_n, clean_urls=not args.no_clean)
